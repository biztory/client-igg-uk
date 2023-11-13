import boto3
import os
import io
import pandas as pd
import datetime
from openpyxl import load_workbook
import numpy as np
import json

# Function to process an Excel file, convert it to CSV, and upload to S3
def process_excel_file(file_key, output_bucket, bucket_name):
    s3_client = boto3.client("s3")
    
    try:
        # Get the Excel file from the source S3 bucket
        response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
        excel_file = response['Body'].read()
        
        # Load the Excel workbook and initialize a list to store DataFrames
        workbook = load_workbook(filename=io.BytesIO(excel_file), read_only=True)
        dfs = []
        
        # Iterate through each sheet in the workbook
        for sheet in workbook.sheetnames:
            if 'LMF' in sheet:
                # Read the sheet as a DataFrame, preprocess the data, and add metadata columns
                df = pd.read_excel(io.BytesIO(excel_file), sheet_name=sheet, skiprows=7, header=0, skipfooter=7, keep_default_na=False)
                df.replace('', np.nan, inplace=True)
                df['Folder_Name'] = os.path.dirname(file_key)
                df['File_Name'] = os.path.basename(file_key)
                df['Sheet_Name'] = sheet
                df['AWS_timestamp'] = datetime.datetime.now()
                df.rename(columns={'Unnamed: 1': 'Type of Fund'}, inplace=True)
                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                df.ffill(axis=0, inplace=True)
                dfs.append(df)
        
        # Concatenate DataFrames and create a final DataFrame
        final_df = pd.concat(dfs, ignore_index=True)
        
        # Prepare the processed data as a CSV
        csv_buffer = io.StringIO()
        final_df.to_csv(csv_buffer, index=False)
        
        # Define the S3 key for the output CSV
        output_key = f"{os.path.dirname(file_key).split('/Weekly Report')[0]}/{os.path.basename(file_key).replace('.xlsx', '.csv')}"
        
        # Upload the processed CSV to the output S3 bucket
        s3_client.put_object(Bucket=output_bucket, Key=output_key, Body=csv_buffer.getvalue())
        
        # Print a message indicating the processing of the file
        print(f"Processed file: {file_key}, Output file: {output_key}")
    
    except Exception as e:
        print(f"Error processing file: {file_key}\nError Details: {str(e)}")

# Function to list files in the source S3 bucket and process eligible Excel files
def list_s3_files_in_folder_using_client():
    s3_client = boto3.client("s3")
    bucket_name = "s3-biz-igg-dev-incomming"
    output_bucket = "s3-biz-igg-dev-processed"
    
    try:
        # List objects in the source S3 bucket
        response = s3_client.list_objects_v2(Bucket=bucket_name)
        files = response.get("Contents")
        sub_folders = set()
        
        # Iterate through the files and identify subfolders
        for file in files:
            if ".xlsx" in file["Key"] and "LMF_and_Profile_Funds_Weekly_Report" in file["Key"]:
                subfolder = os.path.dirname(file["Key"])
                sub_folders.add(subfolder)
        
        # Process files within identified subfolders
        for subfolder in sub_folders:
            response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=f"{subfolder}/")
            files_in_subfolder = response.get("Contents")
            
            # Process each eligible Excel file in the subfolder
            for file in files_in_subfolder:
                process_excel_file(file["Key"], output_bucket, bucket_name)
    
    except Exception as e:
        print(f"Error listing and processing files\nError Details: {str(e)}")

# AWS Lambda handler function
def lambda_handler(event, context):
    try:
        list_s3_files_in_folder_using_client()
        return {
            'statusCode': 200,
            'body': json.dumps('Success')
        }
    except Exception as e:
        return {
            'statusCode': 500,
            'body': json.dumps(f'Error: {str(e)}')
        }

