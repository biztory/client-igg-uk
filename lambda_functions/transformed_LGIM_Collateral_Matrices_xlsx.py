import boto3
import os
import io
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import json

bucket_name = "s3-biz-igg-dev-incomming"
output_bucket = "s3-biz-igg-dev-processed"
s3_client = boto3.client("s3")

def list_files_in_bucket(bucket_name):
    try:
        response = s3_client.list_objects_v2(Bucket=bucket_name)
        return response.get("Contents")
    except Exception as e:
        raise Exception(f"Error listing files in bucket '{bucket_name}': {str(e)}")

def process_excel_file(file_key, output_bucket, bucket_name):
    try:
        response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
        excel_file = response['Body'].read()
        workbook = load_workbook(filename=io.BytesIO(excel_file), read_only=False)
        
        dfs = []

        for sheet in workbook.sheetnames:
            if "Collateral Grids" in sheet:
                for skiprow in range(0, 60, 15):
                    df = pd.read_excel(io.BytesIO(excel_file), skiprows=skiprow, header=[1, 2, 3], nrows=10, keep_default_na=False, sheet_name=sheet, index_col=0)
                    names_list = [str(name) for name in df.columns.get_level_values(1)]
                    names_set = set(names_list)

                    df_names = []

                    for names in names_set:
                        if 'Collateral' in names:
                            extracted_name = names.split('(')[1].split(')')[0]
                            filtered_df = df.loc[:, df.columns.get_level_values(1) == names]
                            df_names.append((extracted_name, filtered_df))

                    for name, df in df_names:
                        table_subheader = df.columns.get_level_values(0)[0]
                        table_name = df.columns.get_level_values(1)[0]
                        df = df.iloc[:, :-1]
                        df.columns = df.columns.droplevel()
                        df.columns = df.columns.droplevel()
                        column_names = ["Interest rate / Inflation shock", 0, 25, 50, 75, 100, 150, 200, 250, 300, 400]
                        df = df[df.columns.intersection(column_names)]
                        df['Table_Subheader'] = table_subheader
                        df['Table_Name'] = table_name
                        df['File_Name'] = os.path.basename(file_key)
                        df['Folder_Name'] = os.path.dirname(file_key)
                        df['AWS_Timestamp'] = datetime.now()
                        df['Interest rate'] = df.index.name = 'Interest rate / Inflation shock'
                        df = df.iloc[0:1]
                        dfs.append(df)
        
        final_df = pd.concat(dfs)

        final_df = final_df.reset_index()
        
        final_df.index.name = "Key"
        
        # Save the CSV data to an in-memory buffer
        csv_buffer = io.BytesIO()
        final_df.to_csv(csv_buffer, index=True)
        csv_buffer.seek(0)  # Reset the buffer's position to the beginning
        s3_key = f"{os.path.dirname(file_key)}/combined_data-{os.path.basename(file_key).replace('.xlsx','.csv')}"
        s3_client.put_object(Bucket=output_bucket, Key=s3_key, Body=csv_buffer.getvalue())
        print(f"'{os.path.basename(file_key)}' has been written to '{s3_key}'")
    except Exception as e:
        raise Exception(f"Error processing Excel file '{file_key}': {str(e)}")

def executing_lgim_coll_matrix():
    try:
        # List objects in the input bucket
        files = list_files_in_bucket(bucket_name=bucket_name)

        sub_folders = set()

        for file in files:
            if "Fund Collateral" in file["Key"]:
                subfolder = os.path.dirname(file["Key"])
                sub_folders.add(subfolder)

        for subfolder in sub_folders:
            response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=f"{subfolder}/")
            files_in_subfolder = response.get("Contents")

            for file in files_in_subfolder:
                if ".xlsx" in file["Key"] and "Fund Collateral" in file["Key"]:
                    try:
                        process_excel_file(file_key=file["Key"], output_bucket=output_bucket, bucket_name=bucket_name)
                    except Exception as e:
                        print(f"Error processing '{file['Key']}': {str(e)}")
    except Exception as e:
        raise Exception(f"Error executing lgim_coll_matrix: {str(e)}")

def lambda_handler(event, context):
    try:
        executing_lgim_coll_matrix()
        return {
            'statusCode': 200,
            'body': json.dumps('Success')
        }
    except Exception as e:
        return {
            'statusCode': 500,
            'body': json.dumps(f'Error: {str(e)}')
        }
